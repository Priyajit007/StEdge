#!/usr/bin/env python3
"""Scan every per-source-file synthesis log under syn/designs/<design>/<module>/,
build an Excel workbook + CSV, and plot within-design and cross-design
comparisons.

Layout produced by synth.py:
    designs/<design>/<module>/{area,areahier,power,powerhier,timing}.log
                              test.log, dc_<module>.tcl, <module>.mapped.v

Here a "design" is one of sobel_stochastic / prewitt_stochastic /
sobel_traditional / prewitt_traditional (used as the *category*) and each row is
one RTL source file ("module"). The eight peripheral modules are byte-identical
across all four designs; only the compute core differs (stochastic sobel_stoch
vs deterministic sobel / prewitt). There are no SRAM hard macros in this RTL, so
(unlike the RWT/MATMUL reference flow) there is no macro-area split.

Metrics parsed per source file:
    area.log    -> combinational / noncombinational / buf-inv / total area (um^2)
                   + cell / comb-cell / seq-cell counts
    power.log   -> internal / switching / dynamic / leakage power (-> uW)
    timing.log  -> worst slack (ps), met?, critical-path arrival (ps),
                   derived achievable period & est. Fmax (1 GHz target)
    test.log    -> SYNTH_OK?, #errors, warning-code histogram
    dc_*.tcl    -> detected clock ports (informational)

Outputs:
    results/synthesis_results.xlsx   (Summary_All + one sheet per design)
    results/summary.csv
    plots/within_<design>.png        per-module area/power/slack within a design
    plots/cross_*.png                between-design comparisons
    plots/cross_compute_core.png     stochastic-vs-traditional compute-core bars
"""
import csv
import os
import re
from collections import defaultdict, OrderedDict

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DESIGNS_DIR = os.path.join(HERE, "designs")
RESULTS_DIR = os.path.join(HERE, "results")
PLOTS_DIR = os.path.join(HERE, "plots")

CLK_PERIOD_PS = 1000.0          # 1 GHz target

# Designs are grouped by operator so the stochastic/traditional pair for each
# edge detector sits adjacent in tables, sheets and the average plots.
CATEGORY_ORDER = ["sobel_stochastic", "sobel_traditional",
                  "prewitt_stochastic", "prewitt_traditional"]

# functional order for the x-axis of the within-design plots (compute cores last)
MODULE_ORDER = ["address_mux", "data_MUX", "camera_reset_clk", "capture_starter",
                "capture", "frame_buffer", "read_write_controller",
                "vga_display", "sobel_stoch", "sobel", "prewitt"]

# paired colours: dark = stochastic, light = traditional, per operator
CAT_COLORS = {"sobel_stochastic": "#4C72B0", "sobel_traditional": "#A6C8E8",
              "prewitt_stochastic": "#DD8452", "prewitt_traditional": "#F3C6A5"}

# the one module that differs between designs (the 8 peripherals are identical);
# used for the direct stochastic-vs-traditional compute-core comparison plot.
COMPUTE_CORE = {"sobel_stochastic": "sobel_stoch",
                "prewitt_stochastic": "sobel_stoch",
                "sobel_traditional": "sobel",
                "prewitt_traditional": "prewitt"}
CORE_ORDER = ["sobel_stochastic", "sobel_traditional",
              "prewitt_stochastic", "prewitt_traditional"]
CORE_LABEL = {"sobel_stochastic": "Sobel\n(stochastic)",
              "sobel_traditional": "Sobel\n(traditional)",
              "prewitt_stochastic": "Prewitt\n(stochastic)",
              "prewitt_traditional": "Prewitt\n(traditional)"}


# ── unit helpers ──────────────────────────────────────────────────────────
def to_uW(val, unit):
    u = unit.strip().lower()
    return {"w": 1e6, "mw": 1e3, "uw": 1.0, "nw": 1e-3,
            "pw": 1e-6, "fw": 1e-9}.get(u, 1.0) * val


# ── parsers ───────────────────────────────────────────────────────────────
def parse_area(path):
    t = open(path, errors="ignore").read()
    out = {}
    pats = {
        "comb_area": r"Combinational area:\s+([\d.]+)",
        "noncomb_area": r"Noncombinational area:\s+([\d.]+)",
        "bufinv_area": r"Buf/Inv area:\s+([\d.]+)",
        "total_cell_area": r"Total cell area:\s+([\d.]+)",
        "num_cells": r"Number of cells:\s+(\d+)",
        "num_comb_cells": r"Number of combinational cells:\s+(\d+)",
        "num_seq_cells": r"Number of sequential cells:\s+(\d+)",
    }
    for k, p in pats.items():
        m = re.search(p, t)
        out[k] = float(m.group(1)) if m else 0.0
    return out


def parse_power(path):
    t = open(path, errors="ignore").read()
    out = {}
    for label, key in [("Cell Internal Power", "internal_uW"),
                       ("Net Switching Power", "switching_uW"),
                       ("Total Dynamic Power", "dynamic_uW"),
                       ("Cell Leakage Power", "leakage_uW")]:
        m = re.search(re.escape(label) + r"\s*=\s*([\d.eE+-]+)\s*(\w+)", t)
        out[key] = to_uW(float(m.group(1)), m.group(2)) if m else 0.0
    out["total_power_uW"] = out["dynamic_uW"] + out["leakage_uW"]
    return out


def parse_timing(path):
    t = open(path, errors="ignore").read()
    arr = re.search(r"data arrival time\s+(-?[\d.]+)", t)
    crit = float(arr.group(1)) if arr else None
    m = re.search(r"slack\s*\((MET|VIOLATED)\)\s+(-?[\d.]+)", t)
    if m:
        met = (m.group(1) == "MET")
        slack = float(m.group(2))
    else:
        met, slack = None, None
    ach_period = (CLK_PERIOD_PS - slack) if slack is not None else None
    fmax = (1000.0 / ach_period) if ach_period and ach_period > 0 else None  # GHz
    return {"worst_slack_ps": slack, "slack_met": met,
            "crit_path_ps": crit, "ach_period_ps": ach_period,
            "est_fmax_ghz": fmax}


def parse_clocks(tcl_path):
    if not os.path.exists(tcl_path):
        return ""
    t = open(tcl_path, errors="ignore").read()
    clks = re.findall(r"create_clock -name (\w+)", t)
    clks = [c for c in clks if c != "vclk"]
    return ",".join(clks) if clks else "comb"


def scan_testlog(path):
    if not os.path.exists(path):
        return {"synth_ok": False, "n_errors": 0, "n_warnings": 0,
                "warn_codes": {}, "errors": []}
    t = open(path, errors="ignore").read()
    errs = re.findall(r"^Error:.*$", t, re.MULTILINE)
    warn_codes = defaultdict(int)
    for code in re.findall(r"Warning:.*?\(([A-Z]+-\d+)\)", t):
        warn_codes[code] += 1
    timed_out = "SYNTH TIMEOUT" in t
    return {"synth_ok": "SYNTH_OK" in t,
            "n_errors": len(errs),
            "n_warnings": sum(warn_codes.values()),
            "warn_codes": dict(warn_codes),
            "timed_out": timed_out,
            "errors": errs}


def module_key(name):
    return MODULE_ORDER.index(name) if name in MODULE_ORDER else 99


def cat_sort_key(r):
    ci = (CATEGORY_ORDER.index(r["category"])
          if r["category"] in CATEGORY_ORDER else 99)
    return (ci, module_key(r["module"]))


# ── collect everything ────────────────────────────────────────────────────
COLUMNS = OrderedDict([
    ("module", "Source file (module)"),
    ("category", "Design"),
    ("clocks", "Clocks"),
    ("status", "Status"),
    ("total_cell_area", "Total cell area (um2)"),
    ("comb_area", "Comb area (um2)"),
    ("noncomb_area", "Noncomb area (um2)"),
    ("bufinv_area", "Buf/Inv area (um2)"),
    ("num_cells", "# cells"),
    ("num_comb_cells", "# comb cells"),
    ("num_seq_cells", "# seq cells"),
    ("dynamic_uW", "Dyn power (uW)"),
    ("leakage_uW", "Leak power (uW)"),
    ("total_power_uW", "Total power (uW)"),
    ("worst_slack_ps", "Worst slack (ps)"),
    ("slack_met", "Slack met"),
    ("crit_path_ps", "Crit path (ps)"),
    ("est_fmax_ghz", "Est Fmax (GHz)"),
    ("n_warnings", "# warnings"),
    ("n_errors", "# errors"),
    ("note", "Note"),
])


def collect():
    rows = []
    for design in sorted(os.listdir(DESIGNS_DIR)):
        ddir = os.path.join(DESIGNS_DIR, design)
        if not os.path.isdir(ddir):
            continue
        for module in sorted(os.listdir(ddir)):
            wd = os.path.join(ddir, module)
            if not os.path.isdir(wd):
                continue
            tl = scan_testlog(os.path.join(wd, "test.log"))
            status = "OK" if tl["synth_ok"] else (
                "TIMEOUT" if tl.get("timed_out") else "FAIL")
            row = {"module": module, "category": design, "status": status,
                   "clocks": parse_clocks(os.path.join(wd, f"dc_{module}.tcl")),
                   "n_warnings": tl["n_warnings"], "n_errors": tl["n_errors"],
                   "warn_codes": tl["warn_codes"]}
            for fn, fn_parse in [("area.log", parse_area),
                                 ("power.log", parse_power),
                                 ("timing.log", parse_timing)]:
                p = os.path.join(wd, fn)
                if os.path.exists(p):
                    row.update(fn_parse(p))
            note = ""
            if status == "OK" and row.get("num_seq_cells", 0) == 0:
                note = "combinational (no registers)"
            row["note"] = note
            rows.append(row)
    rows.sort(key=cat_sort_key)
    return rows


# ── Excel ─────────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(color="FFFFFF", bold=True)
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
FAIL_FILL = PatternFill("solid", fgColor="F8CBAD")
TMO_FILL = PatternFill("solid", fgColor="FFF2CC")


def _fmt(v):
    if isinstance(v, bool):
        return "Yes" if v else "No"
    if isinstance(v, float):
        return round(v, 4)
    return v


def write_sheet(ws, rows):
    keys = list(COLUMNS.keys())
    ws.append([COLUMNS[k] for k in keys])
    for ci in range(1, len(keys) + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    for r in rows:
        ws.append([_fmt(r.get(k, "")) for k in keys])
        rr = ws.max_row
        fill = {"OK": OK_FILL, "TIMEOUT": TMO_FILL}.get(r["status"], FAIL_FILL)
        ws.cell(row=rr, column=4).fill = fill
    ws.freeze_panes = "A2"
    for ci, k in enumerate(keys, 1):
        width = max(len(COLUMNS[k]) + 2,
                    *(len(str(_fmt(r.get(k, "")))) + 2 for r in rows)) \
            if rows else len(COLUMNS[k]) + 2
        ws.column_dimensions[get_column_letter(ci)].width = min(width, 26)


def write_excel(rows):
    os.makedirs(RESULTS_DIR, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary_All"
    write_sheet(ws, rows)
    by_cat = defaultdict(list)
    for r in rows:
        by_cat[r["category"]].append(r)
    for cat in CATEGORY_ORDER:
        if cat in by_cat:
            write_sheet(wb.create_sheet(cat), by_cat[cat])
    out = os.path.join(RESULTS_DIR, "synthesis_results.xlsx")
    wb.save(out)
    print(f"  wrote {out}  ({len(by_cat)} design sheets)")
    csv_keys = [k for k in COLUMNS]
    with open(os.path.join(RESULTS_DIR, "summary.csv"), "w", newline="") as fh:
        w = csv.writer(fh)
        w.writerow([COLUMNS[k] for k in csv_keys])
        for r in rows:
            w.writerow([_fmt(r.get(k, "")) for k in csv_keys])
    print(f"  wrote {os.path.join(RESULTS_DIR, 'summary.csv')}")


# ── plots ─────────────────────────────────────────────────────────────────
def _ok(rows):
    return [r for r in rows if r["status"] == "OK"]


def plot_within_design(rows):
    by_cat = defaultdict(list)
    for r in _ok(rows):
        by_cat[r["category"]].append(r)
    for cat, rs in by_cat.items():
        rs = sorted(rs, key=lambda r: module_key(r["module"]))
        names = [r["module"] for r in rs]
        x = np.arange(len(rs))
        fig, axes = plt.subplots(1, 3, figsize=(min(5 + 1.2 * len(rs), 22), 4.8))
        comb = np.array([r.get("comb_area", 0) for r in rs])
        ncomb = np.array([r.get("noncomb_area", 0) for r in rs])
        axes[0].bar(x, comb, label="Comb logic", color="#4C72B0")
        axes[0].bar(x, ncomb, bottom=comb, label="Noncomb (reg) logic",
                    color="#DD8452")
        axes[0].set_title(f"{cat}: cell area (um2)")
        axes[0].set_ylabel("Area (um2)")
        axes[0].legend(fontsize=8)
        dyn = [r.get("dynamic_uW", 0) for r in rs]
        leak = [r.get("leakage_uW", 0) for r in rs]
        axes[1].bar(x, dyn, label="Dynamic", color="#55A868")
        axes[1].bar(x, leak, bottom=dyn, label="Leakage", color="#C44E52")
        axes[1].set_title(f"{cat}: power (uW)")
        axes[1].set_ylabel("Power (uW)")
        axes[1].legend(fontsize=8)
        slk = [r.get("worst_slack_ps", 0) or 0 for r in rs]
        axes[2].bar(x, slk, color="#8172B3")
        axes[2].axhline(0, color="black", lw=1, ls=":")
        axes[2].set_title(f"{cat}: worst slack (ps) @1GHz")
        axes[2].set_ylabel("Slack (ps)")
        for ax in axes:
            ax.set_xticks(x)
            ax.set_xticklabels(names, rotation=35, ha="right", fontsize=8)
            ax.grid(axis="y", alpha=0.3)
        fig.suptitle(f"Within-design comparison — {cat}", fontweight="bold")
        fig.tight_layout(rect=[0, 0, 1, 0.95])
        fig.savefig(os.path.join(PLOTS_DIR, f"within_{cat}.png"), dpi=140)
        plt.close(fig)


def plot_cross_design(rows):
    rs = _ok(rows)
    if not rs:
        return
    # 1) all modules, area grouped by design (grouped bars per module)
    cats = [c for c in CATEGORY_ORDER if any(r["category"] == c for r in rs)]
    n = max(len(cats), 1)
    modules = [m for m in MODULE_ORDER
               if any(r["module"] == m for r in rs)]
    x = np.arange(len(modules))
    w = 0.8 / n
    fig, ax = plt.subplots(figsize=(min(7 + 1.1 * len(modules), 22), 5.5))
    for i, cat in enumerate(cats):
        vals = []
        for m in modules:
            cell = [r for r in rs if r["module"] == m and r["category"] == cat]
            vals.append(cell[0].get("total_cell_area", 0) if cell else 0)
        ax.bar(x + (i - (n - 1) / 2) * w, vals, w, label=cat,
               color=CAT_COLORS.get(cat, "#888"))
    ax.set_xticks(x)
    ax.set_xticklabels(modules, rotation=35, ha="right", fontsize=8)
    ax.set_ylabel("Total cell area (um2)")
    ax.set_title("Per-source-file area by design "
                 "(peripherals identical; compute core differs)")
    ax.legend(fontsize=8)
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    fig.savefig(os.path.join(PLOTS_DIR, "cross_area_by_module.png"), dpi=140)
    plt.close(fig)

    # 2) design averages: area, power, fmax
    cats = [c for c in CATEGORY_ORDER if any(r["category"] == c for r in rs)]

    def cavg(cat, key):
        vals = [r.get(key, 0) or 0 for r in rs if r["category"] == cat]
        return np.mean(vals) if vals else 0
    metrics = [("total_cell_area", "Avg cell area (um2)"),
               ("total_power_uW", "Avg total power (uW)"),
               ("est_fmax_ghz", "Avg est. Fmax (GHz)")]
    fig, axes = plt.subplots(1, 3, figsize=(14, 5))
    xc = np.arange(len(cats))
    for ax, (key, title) in zip(axes, metrics):
        ax.bar(xc, [cavg(c, key) for c in cats],
               color=[CAT_COLORS.get(c, "#888") for c in cats])
        ax.set_xticks(xc)
        ax.set_xticklabels(cats, rotation=20, ha="right", fontsize=9)
        ax.set_title(title)
        ax.grid(axis="y", alpha=0.3)
    fig.suptitle("Between-design averages", fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.95])
    fig.savefig(os.path.join(PLOTS_DIR, "cross_design_avg.png"), dpi=140)
    plt.close(fig)

    # 3) area vs power scatter, colored by design
    fig, ax = plt.subplots(figsize=(8, 6.5))
    for c in cats:
        cr = [r for r in rs if r["category"] == c]
        ax.scatter([r.get("total_cell_area", 0) for r in cr],
                   [r.get("total_power_uW", 0) for r in cr],
                   label=c, color=CAT_COLORS.get(c, "#888"), s=60,
                   edgecolor="black", linewidth=0.4)
    ax.set_xlabel("Total cell area (um2)")
    ax.set_ylabel("Total power (uW)")
    ax.set_title("Cell area vs Total power (all source files)")
    ax.legend(fontsize=9)
    ax.grid(alpha=0.3)
    fig.tight_layout()
    fig.savefig(os.path.join(PLOTS_DIR, "cross_area_vs_power.png"), dpi=140)
    plt.close(fig)


def plot_compute_core(rows):
    """Direct stochastic-vs-traditional comparison of the compute cores -- the
    only module that differs between the four designs (the eight peripherals are
    byte-identical). Bars: total cell area, total power, sequential-cell (FF)
    count and worst slack, grouped Sobel then Prewitt."""
    pick = {}
    for r in rows:
        core = COMPUTE_CORE.get(r["category"])
        if core and r["module"] == core and r["status"] == "OK":
            pick[r["category"]] = r
    cats = [c for c in CORE_ORDER if c in pick]
    if not cats:
        return
    labels = [CORE_LABEL[c] for c in cats]
    colors = [CAT_COLORS.get(c, "#888") for c in cats]
    x = np.arange(len(cats))
    metrics = [("total_cell_area", "Total cell area (um2)"),
               ("total_power_uW", "Total power (uW)"),
               ("num_seq_cells", "# sequential cells (FF)"),
               ("worst_slack_ps", "Worst slack (ps) @1GHz")]
    fig, axes = plt.subplots(1, 4, figsize=(17, 5))
    for ax, (key, title) in zip(axes, metrics):
        vals = [pick[c].get(key, 0) or 0 for c in cats]
        bars = ax.bar(x, vals, color=colors, edgecolor="black", linewidth=0.4)
        ax.set_xticks(x)
        ax.set_xticklabels(labels, fontsize=8)
        ax.set_title(title)
        ax.grid(axis="y", alpha=0.3)
        if key == "worst_slack_ps":
            ax.axhline(0, color="black", lw=1, ls=":")
        for b, v in zip(bars, vals):
            ax.annotate(f"{v:.1f}" if abs(v) >= 1 else f"{v:.2f}",
                        (b.get_x() + b.get_width() / 2, v),
                        ha="center", va="bottom", fontsize=7)
    fig.suptitle("Compute-core comparison — stochastic vs traditional",
                 fontweight="bold")
    fig.tight_layout(rect=[0, 0, 1, 0.95])
    fig.savefig(os.path.join(PLOTS_DIR, "cross_compute_core.png"), dpi=140)
    plt.close(fig)


def main():
    os.makedirs(PLOTS_DIR, exist_ok=True)
    rows = collect()
    n_ok = sum(1 for r in rows if r["status"] == "OK")
    print(f"Collected {len(rows)} source files ({n_ok} OK)")
    write_excel(rows)
    plot_within_design(rows)
    plot_cross_design(rows)
    plot_compute_core(rows)
    print(f"  plots -> {PLOTS_DIR}/")
    allwarn = defaultdict(int)
    for r in rows:
        for c, n in r.get("warn_codes", {}).items():
            allwarn[c] += n
    if allwarn:
        print("\nTop warning codes across all source files:")
        for c, n in sorted(allwarn.items(), key=lambda x: -x[1])[:15]:
            print(f"  {c:12} {n}")
    bad = [f"{r['category']}/{r['module']}" for r in rows if r["status"] != "OK"]
    if bad:
        print("\nFAILED / TIMEOUT:", ", ".join(bad))


if __name__ == "__main__":
    main()
